import sys
import os
import json

# --- 1. ライブラリパス強制追加 ---
user_lib_path = r"C:\Users\OWNER\AppData\Roaming\Python\Python314\site-packages"
if os.path.exists(user_lib_path) and user_lib_path not in sys.path:
    sys.path.insert(0, user_lib_path)

# --- 2. インポート ---
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
    from openpyxl.drawing.image import Image
    HAS_LIBS = True
except ImportError:
    HAS_LIBS = False

def main():
    if not HAS_LIBS or len(sys.argv) < 3:
        return

    json_input_path = sys.argv[1]
    excel_output_path = sys.argv[2]

    try:
        with open(json_input_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "メモ詳細"

    # --- スタイル定義 ---
    thin = Side(style='thin', color="000000")
    border_all = Border(top=thin, left=thin, right=thin, bottom=thin)
    fill_label = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    font_bold = Font(name='BIZ UDPゴシック', bold=True, size=11)
    font_normal = Font(name='BIZ UDPゴシック', size=10)

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 15

    # 1. タイトル
    ws.merge_cells('A1:E1')
    ws['A1'] = "メモ帳エクスポート報告書"
    ws['A1'].font = Font(name='BIZ UDPゴシック', bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # 2. 基本情報
    ws['A2'] = "管理ID"; ws['B2'] = data.get('id', '---')
    ws['A3'] = "作成日時"; ws['B3'] = data.get('created_at', '')

    for cell in ['A2', 'A3']:
        ws[cell].fill = fill_label
        ws[cell].font = font_bold
        ws[cell].border = border_all
    
    # 基本情報の結合（B-E列）
    ws.merge_cells('B2:E2')
    ws.merge_cells('B3:E3')
    for r in [2, 3]:
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = border_all

    # 3. 内容セクション（動的計算）
    ws['A4'] = "内容"; ws['A4'].fill = fill_label; ws['A4'].font = font_bold; ws['A4'].border = border_all
    ws.merge_cells('A4:E4')

    content = data.get('content', '')
    text_len = len(content)
    line_count = content.count('\n') + 1
    char_rows = (text_len // 40) + (1 if text_len % 40 > 0 else 0)
    
    needed_rows = line_count if line_count > char_rows else char_rows
    if needed_rows < 11: needed_rows = 11
    
    start_row = 5
    end_row = start_row + (needed_rows - 1)
    
    # ★修正：古いopenpyxlでも動くように範囲を文字列で指定
    ws.merge_cells(range_string="A%d:E%d" % (start_row, end_row))
    
    ws.cell(row=start_row, column=1).value = content
    ws.cell(row=start_row, column=1).alignment = Alignment(wrapText=True, vertical='top')
    ws.cell(row=start_row, column=1).font = font_normal

    for r in range(start_row, end_row + 1):
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = border_all

    # 4. 画像セクション
    img_path = data.get('image_path')
    if img_path and os.path.exists(img_path):
        cur_row = end_row + 1
        ws.cell(row=cur_row, column=1).value = "添付画像"
        ws.cell(row=cur_row, column=1).fill = fill_label
        ws.cell(row=cur_row, column=1).font = font_bold
        ws.merge_cells(range_string="A%d:E%d" % (cur_row, cur_row))
        for c in range(1, 6):
            ws.cell(row=cur_row, column=c).border = border_all
        
        try:
            img = Image(img_path)
            scale = 500.0 / img.width
            img.width = 500
            img.height = int(img.height * scale)
            ws.add_image(img, "A%d" % (cur_row + 1))
        except:
            ws.cell(row=cur_row + 1, column=1).value = "(画像の読み込みに失敗しました)"

    # 5. 保存
    wb.save(excel_output_path)

if __name__ == "__main__":
    main()